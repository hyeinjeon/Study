#!/usr/bin/python3
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(filename = "student.xlsx")
sheet_ranges = wb["Sheet1"]

row_num = 2
total_list = []
while True:
        col_midterm = sheet_ranges.cell(row = row_num,column=3).value
        col_final = sheet_ranges.cell(row = row_num,column=4).value
        col_homework = sheet_ranges.cell(row = row_num,column=5).value
        col_attendance = sheet_ranges.cell(row = row_num,column=6).value

        if col_midterm == None:
            break

        col_total = col_midterm * 0.3 + col_final * 0.35 + col_homework * 0.34 + col_attendance
        sheet_ranges.cell(row=row_num,column=7,value=col_total)

        total_list.append(col_total)
        row_num +=1

print(total_list)

student_cnt = row_num - 2
row_num = 2

for i in range(len(total_list)):
    rank = 0
    for j in range(len(total_list)):
        if(total_list[i] <= total_list[j]):
            rank += 1
    print(rank)

    grade_percentage = rank/student_cnt * 100.0
    print(grade_percentage)
    col_total = sheet_ranges.cell(row=row_num, column=7).value
    
    if int(col_total) < 40:
        grade = "F"
    elif grade_percentage <= 15:
        grade = "A+"
    elif grade_percentage <= 30:
        grade = "A"
    elif grade_percentage <= 50:
        grade = "B+"
    elif grade_percentage <= 70:
        grade  = "B"
    elif grade_percentage <= 85:
        grade = "C+"
    else:
        grade = "C"

    sheet_ranges.cell(row=row_num, column=8, value=grade)
    row_num += 1

wb.save(filename="student.xlsx")
